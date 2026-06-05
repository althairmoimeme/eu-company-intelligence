"""Export endpoints — Excel pour listes, CSV enrichi."""
import io
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from ..settings import get_settings
from scraper.db.session import get_session_factory
from scraper.db.models import (
    FavoritesList, FavoritesListItem, Company, Director,
    FounderIntelligence, FinancialHistory
)

router = APIRouter(prefix="/favorites", tags=["export"])
logger = logging.getLogger(__name__)


def _get_db_dep():
    settings = get_settings()
    factory = get_session_factory(settings.DATABASE_PATH)
    async def dep():
        async with factory() as session:
            yield session
    return dep

db_dep = _get_db_dep()

CRM_STATUS_LABELS = {
    "prospect": "Prospect",
    "to_contact": "À contacter",
    "contacted": "Contacté",
    "replied": "A répondu",
    "meeting": "RDV fixé",
    "negociation": "Négociation",
    "deal": "Deal signé",
    "pass": "Passé",
}

OPERATOR_LABELS = {
    "patrimonial": "Patrimonial",
    "builder": "Builder",
    "operator": "Opérateur",
    "founder": "Fondateur",
    "disengaged": "Désengagé",
    "unknown": "Inconnu",
}


@router.get("/lists/{list_id}/export/excel")
async def export_list_excel(
    list_id: int,
    session: AsyncSession = Depends(db_dep),
):
    """Exporte une liste de suivi au format Excel avec toutes les colonnes CRM + FI + financières."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    lst = await session.get(FavoritesList, list_id)
    if not lst:
        raise HTTPException(404, "List not found")

    # Charger tous les items avec relations
    result = await session.execute(
        select(FavoritesListItem)
        .where(FavoritesListItem.list_id == list_id)
        .options(
            selectinload(FavoritesListItem.company)
            .selectinload(Company.directors),
        )
        .order_by(FavoritesListItem.added_at.desc())
    )
    items = result.scalars().all()

    # Charger FI pour toutes les entreprises
    company_ids = [i.company_id for i in items]
    fi_map = {}
    if company_ids:
        fi_result = await session.execute(
            select(FounderIntelligence).where(FounderIntelligence.company_id.in_(company_ids))
        )
        for fi in fi_result.scalars().all():
            fi_map[fi.company_id] = fi

    # Charger derniers snapshots financiers
    fh_map = {}
    if company_ids:
        fh_result = await session.execute(
            select(FinancialHistory)
            .where(FinancialHistory.company_id.in_(company_ids))
            .order_by(FinancialHistory.year.desc())
        )
        for fh in fh_result.scalars().all():
            if fh.company_id not in fh_map:  # garde le plus récent
                fh_map[fh.company_id] = fh

    # ── Créer le classeur ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lst.name[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    signal_fills = {
        "high":     PatternFill("solid", fgColor="FEE2E2"),
        "moderate": PatternFill("solid", fgColor="FEF3C7"),
        "low":      PatternFill("solid", fgColor="F9FAFB"),
    }
    status_fills = {
        "deal":        PatternFill("solid", fgColor="D1FAE5"),
        "meeting":     PatternFill("solid", fgColor="FFEDD5"),
        "negociation": PatternFill("solid", fgColor="EDE9FE"),
        "contacted":   PatternFill("solid", fgColor="EEF2FF"),
        "replied":     PatternFill("solid", fgColor="FFFBEB"),
        "pass":        PatternFill("solid", fgColor="FEE2E2"),
    }

    # En-têtes
    headers = [
        "Entreprise", "Pays", "Secteur", "CA (€)", "Effectifs",
        "Créée le", "Ville",
        # CRM
        "Statut CRM", "Contacté le", "Prochaine action", "Date relance", "Notes",
        # FI
        "Dirigeant", "Rôle", "Âge", "Ancienneté poste",
        "Profil", "Signal vendeur", "Raison signal",
        "Hypothèse why now", "Angle d'approche",
        # Financier
        "Année (fin.)", "CA réel (€)", "Résultat net (€)", "EBITDA (€)",
        # Contact
        "Email", "Tél", "LinkedIn", "Site web",
        # Meta
        "Ajouté le",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 30

    # Données
    for row_idx, item in enumerate(items, 2):
        c = item.company
        fi = fi_map.get(item.company_id)
        fh = fh_map.get(item.company_id)

        # Dirigeant principal
        main_dir = None
        if c.directors:
            def score_d(d):
                s = 0
                if d.birth_year: s += (2026 - d.birth_year)
                role = (d.role or "").lower()
                if any(k in role for k in ["président", "pdg", "gérant", "ceo", "gérant"]): s += 50
                return s
            main_dir = max(c.directors, key=score_d)

        def _fmt_rev(v):
            if v is None: return ""
            return round(v)

        def _fmt_date(d):
            if not d: return ""
            try: return datetime.fromisoformat(str(d)[:10]).strftime("%d/%m/%Y")
            except: return str(d)[:10]

        row_data = [
            c.name,
            c.country,
            c.sector or "",
            _fmt_rev(c.revenue_eur),
            c.employees or "",
            _fmt_date(c.creation_date),
            c.city or "",
            # CRM
            CRM_STATUS_LABELS.get(item.status or "prospect", item.status or ""),
            _fmt_date(item.contacted_at),
            item.next_action or "",
            _fmt_date(item.next_action_date),
            item.notes or "",
            # FI
            (fi.full_name if fi else "") or (main_dir.name if main_dir else ""),
            (fi.current_role if fi else "") or (main_dir.role if main_dir else ""),
            fi.estimated_age if fi else (2026 - main_dir.birth_year if main_dir and main_dir.birth_year else ""),
            fi.years_in_role if fi else "",
            OPERATOR_LABELS.get(fi.operator_type, "") if fi else "",
            fi.seller_signal_strength if fi else "",
            fi.seller_signal_reason if fi else "",
            fi.main_why_now_hypothesis if fi else "",
            fi.recommended_approach_angle if fi else "",
            # Financier
            fh.year if fh else "",
            _fmt_rev(fh.revenue_eur) if fh else "",
            _fmt_rev(fh.net_income_eur) if fh else "",
            _fmt_rev(fh.ebitda_eur) if fh else "",
            # Contact
            (fi.professional_email if fi else "") or c.email or "",
            (fi.phone if fi else "") or c.phone or "",
            fi.linkedin_url if fi else "",
            c.website or "",
            # Meta
            _fmt_date(item.added_at),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = wrap

        # Couleur signal vendeur
        if fi and fi.seller_signal_strength in signal_fills:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = signal_fills[fi.seller_signal_strength]

        # Sur-couche statut CRM pour les colonnes CRM
        status = item.status or "prospect"
        if status in status_fills:
            ws.cell(row=row_idx, column=8).fill = status_fills[status]

    # Largeurs de colonnes
    col_widths = [
        30, 6, 20, 14, 10, 12, 15,
        14, 12, 25, 12, 30,
        25, 20, 6, 12,
        14, 14, 40, 50, 50,
        8, 14, 14, 14,
        25, 15, 35, 25,
        12,
    ]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Métadonnées dans un 2e onglet
    ws2 = wb.create_sheet("Infos liste")
    ws2["A1"] = "Liste"
    ws2["B1"] = lst.name
    ws2["A2"] = "Description"
    ws2["B2"] = lst.description or ""
    ws2["A3"] = "Thèse"
    ws2["B3"] = lst.investment_thesis or ""
    ws2["A4"] = "Exporté le"
    ws2["B4"] = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    ws2["A5"] = "Nb entreprises"
    ws2["B5"] = len(items)

    # Sauvegarder en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = lst.name[:40].replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
